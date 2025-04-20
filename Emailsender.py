import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from openpyxl import load_workbook

# Load the Excel sheet
wb = load_workbook("location of excel file")
sheet = wb.active

# Email settings
sender_email = ""
sender_password = ""
smtp_server = "smtp.gmail.com"  # Use the SMTP server for your email provider
smtp_port = 587

# Connect to the SMTP server
server = smtplib.SMTP(smtp_server, smtp_port)
server.starttls()
server.login(sender_email, sender_password)

# Function to create a personalized ticket
def create_ticket(name):
    ticket_template = f"""
    +---------------------------------------+
    |           Event Ticket                |
    |                                       |
    |  Name: {name}                         |
    |  Event: Your Event Name Here          |
    |  Date: Event Date Here                |
    |                                       |
    |  See you there!                       |
    +---------------------------------------+
    """
    return ticket_template

# Read participant data and send emails
for row in sheet.iter_rows(min_row=2, values_only=True):
    name, email = row
    ticket = create_ticket(name)

    # Create email
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = email
    msg['Subject'] = "Your Event Ticket"
    
    body = MIMEText(ticket, 'plain')
    msg.attach(body)

    # Send the email
    try:
        server.sendmail(sender_email, email, msg.as_string())
        print(f"Ticket sent to {name} at {email}")
    except Exception as e:
        print(f"Failed to send email to {name}: {e}")

# Close the server connection
server.quit()
"""Summary
Purpose of the Script
This Python script automates sending personalized event tickets via email. It reads participant names and email addresses from an Excel file and emails each person a simple text-based event ticket.

How It Works (Step-by-Step)
Import Necessary Libraries

smtplib, email.mime → For sending emails.

openpyxl → For reading Excel files (.xlsx format).

Load Excel File

load_workbook("location of excel file") opens the Excel sheet.

sheet = wb.active uses the first (active) sheet.

It assumes each row has a name and email starting from row 2 (row 1 is typically headers).

Setup Email Details

Sender's email and password (you need to fill these in).

SMTP server setup for Gmail (smtp.gmail.com, port 587).

Establish a connection and login using smtplib.

Create Ticket Function

create_ticket(name) returns a text-based "ticket" with the user's name inserted in a template string.

Read Data & Send Emails

Loops over each row of the Excel sheet.

For each participant:

Creates a personalized ticket.

Prepares the email with subject and body.

Sends the email using the SMTP server.

Prints success or failure message.

Close Server

Ends the SMTP session with server.quit().

📄 Example Excel Layout
Name	Email
Alice Doe	alice@example.com
Bob Smith	bob@example.com
⚠️ Important Notes
You must enable "less secure apps" or use an app password for Gmail to make this work, unless you're using OAuth.
Ensure the Excel file path is correct.
Never hardcode your real email/password in production scripts.
"""
